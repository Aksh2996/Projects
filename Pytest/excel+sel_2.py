import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def update_excel(colname,rowname,new):
    import openpyxl
    book = openpyxl.load_workbook(r"C:\Users\DELL\OneDrive\Desktop\5lv\download.xlsx")
    sheet = book.active
    dics = {}

    for i in range(1, sheet.max_column + 1):
        if (sheet.cell(row=1, column=i).value) == colname:
            dics["colssss"] = i
    for j in range(1, sheet.max_row + 1):
        for k in range(1, sheet.max_column + 1):
            if (sheet.cell(row=j, column=k).value) == rowname:
                dics["rows"] = j
    print(dics["rows"])
    sheet.cell(row=dics["rows"], column=dics["colssss"]).value = new
    book.save(r"C:\Users\DELL\OneDrive\Desktop\5lv\download.xlsx")
newss=110
update_excel("price",109,newss)
driver = webdriver.Chrome()
driver.implicitly_wait(30)
driver.get("https://rahulshettyacademy.com/upload-download-test/")
wait = WebDriverWait(driver, 30)
download_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#downloadButton")))
download_button.click()
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(r"C:\Users\DELL\OneDrive\Desktop\5lv\download.xlsx")
#file_input = wait.until(EC.visibility_of_element_located((By.XPATH, "//div[contains(text(), 'Updated Excel Data')]"))).text
file_input = wait.until(EC.visibility_of_element_located((By.XPATH,"//div[text()='Updated Excel Data Successfully.']"))).text
print(file_input)
value=driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
news=driver.find_element(By.XPATH,"//div[text()='Apple']/parent::div/parent::div/div[@id='cell-"+value+"-undefined']").text
print(news)
assert int(news)==newss

